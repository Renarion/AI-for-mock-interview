"""
Это скрипт, который разово импортирует задачи из excel файлк в таблицу tasks в БД.
Если задачи уже есть в БД, то они не будут импортированы повторно.
Usage:
   сначала загрузить tasks.xlsx в корень сервера, потом выполнить этим команды
   cd ~/AI-for-mock-interview/backend
   python -m task_migrator.import_tasks ~/tasks.xlsx

First row = headers. Map your column names to DB fields below (COLUMN_MAPPING).
Allowed values: company_tier = tier1|tier2; employee_level = junior|middle|senior;
  type = product_analyst|data_analyst; subtype = statistics|ab_testing|probability|python|sql|random.
"""
import asyncio
import csv
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sqlalchemy import select
from openpyxl import load_workbook

from app.database import async_session_maker, init_db
from app.models.task import Task

# Map your sheet column names -> Task model fields.
# Edit keys to match your CSV/Excel header row; values are DB column names.
COLUMN_MAPPING = {
    "task_question": "task_question",
    "task_answer": "task_answer",
    "company_tier": "company_tier",
    "employee_level": "employee_level",
    "type": "type",
    "subtype": "subtype",
    "source": "source",
    # Common alternatives (uncomment or add your headers):
    # "Question": "task_question",
    # "Answer": "task_answer",
    # "Tier": "company_tier",
    # "Level": "employee_level",
    # "Role": "type",
    # "Topic": "subtype",
}

REQUIRED = {"task_question", "company_tier", "employee_level", "type", "subtype"}
OPTIONAL = {"task_answer", "source"}

def _normalize_header(h: str) -> str:
    return (h or "").strip().lower().replace(" ", "_").replace("-", "_")

def _read_csv(path: Path) -> list[dict]:
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        raw_headers = reader.fieldnames or []
        # Map file headers -> db field (case-insensitive match to COLUMN_MAPPING keys)
        header_to_field = {}
        for raw in raw_headers:
            norm = _normalize_header(raw)
            for map_key, db_field in COLUMN_MAPPING.items():
                if _normalize_header(map_key) == norm:
                    header_to_field[raw] = db_field
                    break
        for row in reader:
            rec = {}
            for file_col, db_field in header_to_field.items():
                val = (row.get(file_col) or "").strip()
                if val:
                    rec[db_field] = val
            if rec:
                rows.append(rec)
    return rows

def _read_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        wb.close()
        return []
    raw_headers = [str(h).strip() if h is not None else "" for h in header_row]
    header_to_field = {}
    for i, raw in enumerate(raw_headers):
        norm = _normalize_header(raw)
        for map_key, db_field in COLUMN_MAPPING.items():
            if _normalize_header(map_key) == norm:
                header_to_field[i] = db_field
                break
    rows = []
    for row in rows_iter:
        rec = {}
        for i, db_field in header_to_field.items():
            if i < len(row):
                val = row[i]
                if val is not None and str(val).strip():
                    rec[db_field] = str(val).strip()
        if rec:
            rows.append(rec)
    wb.close()
    return rows


def load_rows(path: Path) -> list[dict]:
    suf = path.suffix.lower()
    if suf == ".csv":
        return _read_csv(path)
    if suf in (".xlsx", ".xls"):
        return _read_xlsx(path)
    raise SystemExit(f"Unsupported format: {suf}. Use .csv or .xlsx")


def normalize_row(row: dict) -> None:
    """Normalize values for DB: lowercase, underscores, empty -> None."""
    if "task_answer" in row and (row["task_answer"] or "").strip() == "":
        row["task_answer"] = None
    if "source" in row and (row["source"] or "").strip() == "":
        row["source"] = None
    for k in ("company_tier", "employee_level", "type", "subtype"):
        if k in row and row[k]:
            row[k] = str(row[k]).strip().lower().replace(" ", "_").replace("-", "_")
    if row.get("company_tier") == "tier_1":
        row["company_tier"] = "tier1"
    elif row.get("company_tier") == "tier_2":
        row["company_tier"] = "tier2"


def validate_row(row: dict, index: int) -> None:
    missing = REQUIRED - set(row.keys())
    if missing:
        raise SystemExit(f"Row {index + 1}: missing required fields: {missing}. Row keys: {list(row.keys())}")
    tier = row.get("company_tier", "")
    if tier not in ("tier1", "tier2"):
        raise SystemExit(f"Row {index + 1}: company_tier must be tier1 or tier2, got: {row.get('company_tier')}")
    level = row.get("employee_level", "")
    if level not in ("junior", "middle", "senior"):
        raise SystemExit(f"Row {index + 1}: employee_level must be junior|middle|senior, got: {row.get('employee_level')}")
    typ = row.get("type", "")
    if typ not in ("product_analyst", "data_analyst"):
        raise SystemExit(f"Row {index + 1}: type must be product_analyst or data_analyst, got: {row.get('type')}")
    subtype = row.get("subtype", "")
    allowed_sub = ("statistics", "ab_testing", "probability", "python", "sql", "random")
    if subtype not in allowed_sub:
        raise SystemExit(f"Row {index + 1}: subtype must be one of {allowed_sub}, got: {row.get('subtype')}")


async def import_tasks(file_path: str) -> None:
    path = Path(file_path).resolve()
    if not path.is_file():
        raise SystemExit(f"File not found: {path}")

    rows = load_rows(path)
    if not rows:
        raise SystemExit("No data rows found (or headers did not match COLUMN_MAPPING).")

    for i, row in enumerate(rows):
        row.setdefault("task_answer", None)
        row.setdefault("source", None)
        normalize_row(row)
        validate_row(row, i)

    await init_db()
    async with async_session_maker() as session:
        # Load existing task_question texts to skip duplicates
        result = await session.execute(select(Task.task_question))
        existing_questions = {row[0] for row in result.fetchall()}

        added = 0
        skipped = 0
        for row in rows:
            if row["task_question"] in existing_questions:
                skipped += 1
                continue
            session.add(Task(**row))
            existing_questions.add(row["task_question"])
            added += 1
        await session.commit()

    print(f"Imported {added} tasks from {path.name}." + (f" Skipped {skipped} duplicates." if skipped else ""))


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python -m task_migrator.import_tasks <path/to/tasks.csv or tasks.xlsx>")
        sys.exit(1)
    asyncio.run(import_tasks(sys.argv[1]))
